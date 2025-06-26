import os
from zipfile import ZipFile

from PyPDF2 import PdfReader
import zipfile
from openpyxl import load_workbook
from paths import RESOURCE_DIR
import csv


def test_pdf():
    zip_file: ZipFile
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, 'files.zip')) as zip_file:
        with zip_file.open('sample.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()

    assert "Home work 07" in text


def test_xlsx():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, 'files.zip')) as zip_file:
        with zip_file.open('sample.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

    assert sheet.cell(row=3, column=8).value == 'Грязьево'


def test_csv():
    with zipfile.ZipFile(os.path.join(RESOURCE_DIR, 'files.zip')) as zip_file:
        with zip_file.open('sample.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            null_row = csvreader[0]

    assert null_row[0] == '№ п/п;Фамилия;Имя;Отчество;Дата рождения;Место рождения;Гражданство;город проживания'




